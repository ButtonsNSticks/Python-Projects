"""Challenge: Write a program to Transpose row & column of the cells in the spreadsheet."""

# Import modules.
from openpyxl import load_workbook
import os
import pandas as pd
import sys

# Define spreadsheet name

sheet = "example.xlsx"

# Open spreadsheet

if os.path.isfile(sheet):
    wb = load_workbook(filename=sheet)
    ws = wb.active
else:
    print("File {} not found".format(sheet))
    sys.exit()

# Grab the max row and max column of active sheet

max_row = ws.max_row
max_col = ws.max_column

# Define dataframe for pandas

df = pd.read_excel(sheet)

# Transpose and export to Excel

df.T.to_excel("transpose.xlsx", index=False)

# Todo: Save the sheet.

