#! python3

"""Challenge: To make a take an integer N from the command line and create an NxN multiplication
Table in an Excel spreadsheet"""

# Import modules

import openpyxl  # Needed to work with spreadsheets
from openpyxl import Workbook
from openpyxl.styles import Font
import sys  # Needed to take in command line arguments

# Define the Font Style to be used with Row and Column header (i.e. BOLD)

ft = Font(bold=True)

# Check that user has entered an integer N from the command line, quit if not.

if len(sys.argv) < 2:
    print("Usage: multiplicationtable.py N  -  Where N is a whole number")
    sys.exit()
else:
    n = int(sys.argv[1])  # The user has entered a value for N, so we save it to variable n.
    filename = "{} by {} Multiplication Grid.xlsx".format(n, n)  # Sets up the name of the XLSX file we will save.

wb = Workbook()  # Create the Excel Spreadsheet
ws = wb.active  # Get the active worksheet

""" First sort out the column & row headers"""
for col in range(n):
    ws.cell(1, col+2).value = col+1
    ws.cell(1, col+2).font = ft

for row in range(n):
    ws.cell(row+2, 1).value = row+1
    ws.cell(row+2, 1).font = ft

""" Now do the N x N grid"""

for row in range(n):
    for col in range (n):
        ws.cell(row+2, col+2).value = int(row+1)*int(col+1)

"""Save the workbook"""

wb.save(filename)
