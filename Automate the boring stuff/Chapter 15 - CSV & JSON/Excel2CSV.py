#! python3
# Excel2CSV.py - Convert an excel file to a CSV file.

import pandas as pd
from os.path import exists
from openpyxl import load_workbook

file_name = "example.xlsx"

if not exists(file_name):  # Check if the file exists

    print("File {} not found  - Exiting now".format(file_name))
    exit()

else:

    wb = load_workbook(file_name)  # Load the spreadsheet.

    # Get a list of all the worksheets in the spreadsheet.

    sheet_list = wb.sheetnames

    for sheet in sheet_list:  # Loop through the sheets

        read_file = pd.read_excel(file_name, sheet_name=sheet)
        csv_output_file = "example - {}.csv".format(sheet)
        read_file.to_csv(csv_output_file, index=None, header=True)
